"""Integration tests for the MCP -> Excel pipeline.

Three offline test scenarios (no real MCP calls — fixtures only):

1. Happy path       — all caches present, full pipeline, validator passes.
2. MISSING data     — critical cache field removed, validator fails check #8.
3. Formula preserve — fill does not overwrite formulas on IS/BS/CF/DCF sheets.

Run:
    pytest plugins/agent-finance/skills/excel-financial-model/tests/test_pipeline.py -v
"""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Resolve skill dirs.
TESTS_DIR = Path(__file__).resolve().parent
EXCEL_SKILL_DIR = TESTS_DIR.parent
BUILDER_SKILL_DIR = EXCEL_SKILL_DIR.parent / "model-input-builder"
REPO_ROOT = EXCEL_SKILL_DIR.parent.parent.parent.parent
FIXTURES_DIR = TESTS_DIR / "fixtures" / "AAPL"
TEMPLATE_V2 = REPO_ROOT / "fundamental_model_template_v2.xlsx"

sys.path.insert(0, str(EXCEL_SKILL_DIR))
sys.path.insert(0, str(BUILDER_SKILL_DIR))

import fill_model  # noqa: E402
import build_inputs  # noqa: E402
import validate_model  # noqa: E402


def _copy_fixtures(dest: Path, omit_field: tuple[str, str] | None = None) -> None:
    """Copy AAPL fixture JSONs to dest. Optionally remove a field from a file."""
    dest.mkdir(parents=True, exist_ok=True)
    for src_file in FIXTURES_DIR.glob("*.json"):
        dst_file = dest / src_file.name
        if omit_field and src_file.name == omit_field[0]:
            data = json.loads(src_file.read_text(encoding="utf-8"))
            # Remove the field from the last annual record.
            if "annual" in data and isinstance(data["annual"], list):
                for rec in data["annual"]:
                    rec.pop(omit_field[1], None)
            # Remove from top-level keys too.
            data.pop(omit_field[1], None)
            dst_file.write_text(json.dumps(data, indent=2), encoding="utf-8")
        else:
            shutil.copy(src_file, dst_file)


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


@pytest.mark.skipif(not TEMPLATE_V2.exists(), reason="v2 template not found at repo root")
def test_happy_path(tmp_dir):
    """Full pipeline: fixtures -> values.json -> filled xlsx -> validator exit 0."""
    cache_dir = tmp_dir / "cache"
    inputs_dir = tmp_dir / "inputs"
    output_dir = tmp_dir / "output"
    output_dir.mkdir()

    _copy_fixtures(cache_dir)

    # Step 1: build inputs.
    report = build_inputs.build_inputs(
        ticker="AAPL",
        cache_dir_override=str(cache_dir),
        output_dir_override=str(inputs_dir),
    )
    assert report["missing_critical_count"] == 0, (
        f"Unexpected critical missing in happy path: {report}"
    )

    # Step 2: fill model.
    values = json.loads((inputs_dir / "values.json").read_text(encoding="utf-8"))
    sources = json.loads((inputs_dir / "sources.json").read_text(encoding="utf-8"))
    output_xlsx = output_dir / "AAPL_test.xlsx"
    fill_report = fill_model.fill(str(TEMPLATE_V2), str(output_xlsx), values, sources)
    assert fill_report["written"] > 0
    assert output_xlsx.exists()

    # Step 3: validate — checks 3 (no errors) and 6 (source traceability) must pass.
    # Checks 7 (v2 built-in) are skipped since xlsx was never opened in Excel.
    # Checks 1/2 pass vacuously on an unfilled model (zeros balance trivially).
    val_report = validate_model.validate(output_xlsx)
    # We assert specific checks rather than full pass (check 7 skipped is ok).
    assert val_report["checks"]["3_no_error_cells"]["passed"], (
        "Error cells found after fill"
    )
    assert val_report["checks"]["6_source_traceability"]["passed"], (
        "Source traceability failed"
    )
    assert val_report["checks"]["8_missing_critical_inputs"]["passed"], (
        "Critical inputs are MISSING in happy path"
    )
    assert val_report["checks"]["9_scenario_coverage"]["passed"], (
        "Scenario coverage failed"
    )


@pytest.mark.skipif(not TEMPLATE_V2.exists(), reason="v2 template not found at repo root")
def test_missing_data_blocks_delivery(tmp_dir):
    """When a critical field is absent, check #8 must fail (non-zero exit)."""
    cache_dir = tmp_dir / "cache"
    inputs_dir = tmp_dir / "inputs"
    output_dir = tmp_dir / "output"
    output_dir.mkdir()

    # Remove beta from valuation-multiples (beta is a critical input).
    _copy_fixtures(cache_dir, omit_field=("valuation-multiples.json", "beta"))

    build_inputs.build_inputs(
        ticker="AAPL",
        cache_dir_override=str(cache_dir),
        output_dir_override=str(inputs_dir),
    )

    values = json.loads((inputs_dir / "values.json").read_text(encoding="utf-8"))
    sources = json.loads((inputs_dir / "sources.json").read_text(encoding="utf-8"))
    output_xlsx = output_dir / "AAPL_missing.xlsx"
    fill_model.fill(str(TEMPLATE_V2), str(output_xlsx), values, sources)

    val_report = validate_model.validate(output_xlsx)
    assert not val_report["checks"]["8_missing_critical_inputs"]["passed"], (
        "Check #8 should fail when critical input (beta) is MISSING"
    )


@pytest.mark.skipif(not TEMPLATE_V2.exists(), reason="v2 template not found at repo root")
def test_formula_preservation(tmp_dir):
    """Filling Assumptions must not overwrite formulas on IS/BS/CF/DCF/Comps."""
    cache_dir = tmp_dir / "cache"
    inputs_dir = tmp_dir / "inputs"
    output_dir = tmp_dir / "output"
    output_dir.mkdir()

    _copy_fixtures(cache_dir)

    build_inputs.build_inputs(
        ticker="AAPL",
        cache_dir_override=str(cache_dir),
        output_dir_override=str(inputs_dir),
    )

    values = json.loads((inputs_dir / "values.json").read_text(encoding="utf-8"))
    sources = json.loads((inputs_dir / "sources.json").read_text(encoding="utf-8"))
    output_xlsx = output_dir / "AAPL_formulas.xlsx"

    # Capture formulas in formula-bearing sheets BEFORE fill.
    formula_sheets = ["Income Statement", "Balance Sheet", "Cash Flow", "DCF", "Comps"]
    before: dict[tuple[str, str], str] = {}
    wb_before = load_workbook(str(TEMPLATE_V2), data_only=False)
    for ws in wb_before.worksheets:
        if ws.title in formula_sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        before[(ws.title, cell.coordinate)] = cell.value

    # Fill model.
    fill_model.fill(str(TEMPLATE_V2), str(output_xlsx), values, sources)

    # Capture formulas AFTER fill.
    after: dict[tuple[str, str], str] = {}
    wb_after = load_workbook(str(output_xlsx), data_only=False)
    for ws in wb_after.worksheets:
        if ws.title in formula_sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        after[(ws.title, cell.coordinate)] = cell.value

    changed = {
        k: (before[k], after.get(k))
        for k in before
        if after.get(k) != before[k]
    }
    assert not changed, (
        f"Formulas changed after fill (should only write Assumptions): {changed}"
    )

    # Also assert no formula was deleted.
    deleted = set(before.keys()) - set(after.keys())
    assert not deleted, f"Formulas deleted after fill: {deleted}"

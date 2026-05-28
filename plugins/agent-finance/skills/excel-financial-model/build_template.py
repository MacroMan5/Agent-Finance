"""Build the 3-statement + DCF + Sensitivity Excel skeleton.

Generates two artefacts that MUST stay in sync, so they are produced from a
single in-memory specification:

    <skill>/template/model_template.xlsx
    <skill>/reference/cell_map.json

The xlsx is a *skeleton*: IS / BS / CF are wired (the three statements close
on each other via formulas), the Assumptions tab carries every input, and
the DCF / Sensitivity tabs have named cells but the formulas are stubbed
out (marked TODO in the body). Phase 4 ships this skeleton and validates it;
DCF / Sensitivity formulas are filled in a later phase.

Run via the plugin's resolved path:

    python "$CLAUDE_PLUGIN_ROOT/skills/excel-financial-model/build_template.py"
"""

from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


SKILL_DIR = Path(__file__).resolve().parent
PLUGIN_ROOT = Path(os.environ.get("CLAUDE_PLUGIN_ROOT", SKILL_DIR.parent.parent))
TEMPLATE_PATH = SKILL_DIR / "template" / "model_template.xlsx"
CELL_MAP_PATH = SKILL_DIR / "reference" / "cell_map.json"

FORECAST_YEARS = 5  # Y1 .. Y5

# ---------------------------------------------------------------------------
# Single source of truth — assumption rows.
# Each row: (logical_name, label, unit, default_value).
# Sources are filled in by fill_model.py at write time.
# ---------------------------------------------------------------------------

ASSUMPTIONS: list[tuple[str, str, str, object]] = [
    # reporting currency (string)
    ("reporting_currency", "Reporting currency", "ISO-4217", "USD"),
    # pricing & shares
    ("share_price", "Share price", "ccy/share", 0.0),
    ("diluted_shares_out", "Diluted shares outstanding", "m shares", 0.0),
    # historical anchor (FY-1)
    ("rev_fy_minus_1", "Revenue (FY-1)", "ccy m", 0.0),
    ("gross_margin_fy_minus_1_pct", "Gross margin (FY-1)", "%", 0.0),
    ("opex_pct_fy_minus_1", "Opex / revenue (FY-1)", "%", 0.0),
    ("da_pct_fy_minus_1", "D&A / revenue (FY-1)", "%", 0.0),
    ("tax_rate_fy_minus_1_pct", "Effective tax rate (FY-1)", "%", 0.0),
    # balance-sheet anchor (FY-1)
    ("cash_fy_minus_1", "Cash & equivalents (FY-1)", "ccy m", 0.0),
    ("debt_fy_minus_1", "Gross debt (FY-1)", "ccy m", 0.0),
    ("equity_fy_minus_1", "Book equity (FY-1)", "ccy m", 0.0),
    ("interest_rate_pct", "Average cost of debt", "%", 0.0),
]

# Per-year forecast driver families: each driver becomes Y1..Y5 rows.
DRIVER_FAMILIES: list[tuple[str, str, str]] = [
    ("rev_growth", "Revenue growth", "%"),
    ("gross_margin", "Gross margin", "%"),
    ("opex_pct", "Opex / revenue", "%"),
    ("da_pct", "D&A / revenue", "%"),
    ("capex_pct", "Capex / revenue", "%"),
    ("nwc_pct_revenue", "NWC / revenue", "%"),
    ("tax_rate", "Effective tax rate", "%"),
]

# DCF + Sensitivity inputs (TODO — formulas stubbed in their sheets).
DCF_INPUTS: list[tuple[str, str, str, float]] = [
    ("wacc_pct", "WACC", "%", 0.0),
    ("terminal_growth_pct", "Terminal growth", "%", 0.0),
]

SENSITIVITY_INPUTS: list[tuple[str, str, str, float]] = [
    ("wacc_min_pct", "Sensitivity WACC min", "%", 0.0),
    ("wacc_max_pct", "Sensitivity WACC max", "%", 0.0),
    ("tg_min_pct", "Sensitivity terminal-growth min", "%", 0.0),
    ("tg_max_pct", "Sensitivity terminal-growth max", "%", 0.0),
]


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin", color="CCCCCC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TODO_FILL = PatternFill("solid", fgColor="F8CBAD")


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_label(cell):
    cell.font = LABEL_FONT


def style_input(cell):
    cell.fill = INPUT_FILL
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_todo(cell):
    cell.fill = TODO_FILL


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build() -> dict:
    wb = Workbook()
    # remove default sheet, add ours in order
    wb.remove(wb.active)

    cell_map: dict[str, dict] = {}

    def register(name: str, sheet: str, cell: str, kind: str, named_range: str | None = None):
        nr = named_range or name
        cell_map[name] = {
            "sheet": sheet,
            "cell": cell,
            "named_range": nr,
            "kind": kind,
        }
        ref = f"'{sheet}'!${cell.rstrip('0123456789')}${cell.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ')}"
        wb.defined_names[nr] = DefinedName(name=nr, attr_text=ref)

    # ---- Assumptions tab -----------------------------------------------------
    ws_a = wb.create_sheet("Assumptions")
    headers = ["name", "label", "value", "unit", "source"]
    for i, h in enumerate(headers, 1):
        c = ws_a.cell(row=1, column=i, value=h)
        style_header(c)
    ws_a.column_dimensions["A"].width = 32
    ws_a.column_dimensions["B"].width = 36
    ws_a.column_dimensions["C"].width = 14
    ws_a.column_dimensions["D"].width = 12
    ws_a.column_dimensions["E"].width = 60

    row = 2

    # static rows
    for name, label, unit, default in ASSUMPTIONS:
        ws_a.cell(row=row, column=1, value=name)
        ws_a.cell(row=row, column=2, value=label)
        v = ws_a.cell(row=row, column=3, value=default)
        style_input(v)
        ws_a.cell(row=row, column=4, value=unit)
        # column E (source) left blank — fill_model.py stamps it.
        register(name, "Assumptions", f"C{row}", "input")
        row += 1

    # per-year drivers
    for fam_key, fam_label, unit in DRIVER_FAMILIES:
        for y in range(1, FORECAST_YEARS + 1):
            name = f"{fam_key}_y{y}_pct" if unit == "%" else f"{fam_key}_y{y}"
            ws_a.cell(row=row, column=1, value=name)
            ws_a.cell(row=row, column=2, value=f"{fam_label} Y{y}")
            v = ws_a.cell(row=row, column=3, value=0.0)
            style_input(v)
            ws_a.cell(row=row, column=4, value=unit)
            register(name, "Assumptions", f"C{row}", "input")
            row += 1

    # DCF inputs (still on Assumptions for single source of truth)
    for name, label, unit, default in DCF_INPUTS:
        ws_a.cell(row=row, column=1, value=name)
        ws_a.cell(row=row, column=2, value=label)
        v = ws_a.cell(row=row, column=3, value=default)
        style_input(v)
        ws_a.cell(row=row, column=4, value=unit)
        register(name, "Assumptions", f"C{row}", "input")
        row += 1

    # Sensitivity inputs
    for name, label, unit, default in SENSITIVITY_INPUTS:
        ws_a.cell(row=row, column=1, value=name)
        ws_a.cell(row=row, column=2, value=label)
        v = ws_a.cell(row=row, column=3, value=default)
        style_input(v)
        ws_a.cell(row=row, column=4, value=unit)
        register(name, "Assumptions", f"C{row}", "input")
        row += 1

    # ---- IS (Income Statement) ----------------------------------------------
    ws_is = wb.create_sheet("IS")
    is_headers = ["Line"] + [f"FY-1"] + [f"Y{y}" for y in range(1, FORECAST_YEARS + 1)]
    for i, h in enumerate(is_headers, 1):
        c = ws_is.cell(row=1, column=i, value=h)
        style_header(c)
    ws_is.column_dimensions["A"].width = 30
    for col in range(2, len(is_headers) + 1):
        ws_is.column_dimensions[get_column_letter(col)].width = 14

    # Row index conventions on IS:
    #   row 2: revenue
    #   row 3: COGS
    #   row 4: gross profit
    #   row 5: opex
    #   row 6: D&A
    #   row 7: EBIT
    #   row 8: interest expense
    #   row 9: pretax income
    #   row 10: tax
    #   row 11: net income
    labels = [
        "Revenue",
        "COGS",
        "Gross profit",
        "Opex",
        "D&A",
        "EBIT",
        "Interest expense",
        "Pretax income",
        "Tax",
        "Net income",
    ]
    for i, lbl in enumerate(labels, start=2):
        c = ws_is.cell(row=i, column=1, value=lbl)
        style_label(c)

    # FY-1 anchor — column B
    ws_is["B2"] = "=rev_fy_minus_1"
    ws_is["B3"] = "=-B2*(1-gross_margin_fy_minus_1_pct/100)"
    ws_is["B4"] = "=B2+B3"
    ws_is["B5"] = "=-B2*opex_pct_fy_minus_1/100"
    ws_is["B6"] = "=-B2*da_pct_fy_minus_1/100"
    ws_is["B7"] = "=B4+B5+B6"
    ws_is["B8"] = "=-debt_fy_minus_1*interest_rate_pct/100"
    ws_is["B9"] = "=B7+B8"
    ws_is["B10"] = "=-MAX(B9,0)*tax_rate_fy_minus_1_pct/100"
    ws_is["B11"] = "=B9+B10"

    # Forecast columns C..G (Y1..Y5)
    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)  # C..G
        prev = get_column_letter(2 + y - 1)
        ws_is[f"{col}2"] = f"={prev}2*(1+rev_growth_y{y}_pct/100)"
        ws_is[f"{col}3"] = f"=-{col}2*(1-gross_margin_y{y}_pct/100)"
        ws_is[f"{col}4"] = f"={col}2+{col}3"
        ws_is[f"{col}5"] = f"=-{col}2*opex_pct_y{y}/100"
        ws_is[f"{col}6"] = f"=-{col}2*da_pct_y{y}/100"
        ws_is[f"{col}7"] = f"={col}4+{col}5+{col}6"
        # interest based on previous-year debt (kept flat for the skeleton)
        ws_is[f"{col}8"] = f"=-debt_fy_minus_1*interest_rate_pct/100"
        ws_is[f"{col}9"] = f"={col}7+{col}8"
        ws_is[f"{col}10"] = f"=-MAX({col}9,0)*tax_rate_y{y}_pct/100"
        ws_is[f"{col}11"] = f"={col}9+{col}10"

    # Register key outputs (Y5 used downstream, plus net income each year)
    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)
        register(f"is_revenue_y{y}", "IS", f"{col}2", "output", f"is_revenue_y{y}")
        register(f"is_ebit_y{y}", "IS", f"{col}7", "output", f"is_ebit_y{y}")
        register(f"is_net_income_y{y}", "IS", f"{col}11", "output", f"is_net_income_y{y}")
        register(f"is_da_y{y}", "IS", f"{col}6", "output", f"is_da_y{y}")

    # ---- BS (Balance Sheet) -------------------------------------------------
    ws_bs = wb.create_sheet("BS")
    for i, h in enumerate(is_headers, 1):  # same column layout as IS
        c = ws_bs.cell(row=1, column=i, value=h)
        style_header(c)
    ws_bs.column_dimensions["A"].width = 30
    for col in range(2, len(is_headers) + 1):
        ws_bs.column_dimensions[get_column_letter(col)].width = 14

    bs_labels = [
        "Cash",                     # 2
        "Net working capital",      # 3
        "Net PP&E (cumulative)",    # 4 — purely book; equals prior + capex - DA
        "Total assets",             # 5
        "Debt",                     # 6
        "Equity",                   # 7
        "Total liab + equity",      # 8
        "Check (Assets - L&E)",     # 9 — must be 0
    ]
    for i, lbl in enumerate(bs_labels, start=2):
        c = ws_bs.cell(row=i, column=1, value=lbl)
        style_label(c)

    # FY-1 anchor
    ws_bs["B2"] = "=cash_fy_minus_1"
    ws_bs["B3"] = "=0"   # NWC anchor (skeleton: assume zero opening NWC; CF reflects deltas)
    ws_bs["B4"] = "=0"   # opening net PP&E book — skeleton starts at zero, drift accrues forward
    ws_bs["B5"] = "=B2+B3+B4"
    ws_bs["B6"] = "=debt_fy_minus_1"
    ws_bs["B7"] = "=equity_fy_minus_1"
    ws_bs["B8"] = "=B6+B7"
    ws_bs["B9"] = "=B5-B8"

    # Forecast: cash is solved by CF; NWC and PPE are derived from drivers;
    # equity rolls with retained earnings.
    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)
        prev = get_column_letter(2 + y - 1)
        # cash: closed by linking to CF tab cumulative cash row (row 6 on CF, see below)
        ws_bs[f"{col}2"] = f"=CF!{col}6"
        # NWC: nwc_pct_revenue * revenue
        ws_bs[f"{col}3"] = f"=nwc_pct_revenue_y{y}/100*IS!{col}2"
        # net PP&E: prior + capex - D&A
        ws_bs[f"{col}4"] = f"={prev}4 + capex_pct_y{y}/100*IS!{col}2 + IS!{col}6"
        # total assets
        ws_bs[f"{col}5"] = f"={col}2+{col}3+{col}4"
        # debt: flat (skeleton)
        ws_bs[f"{col}6"] = f"={prev}6"
        # equity: prior + net income
        ws_bs[f"{col}7"] = f"={prev}7 + IS!{col}11"
        ws_bs[f"{col}8"] = f"={col}6+{col}7"
        ws_bs[f"{col}9"] = f"={col}5-{col}8"

    # Register total-assets and L+E for validator
    for y in range(0, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)
        label = "fy_minus_1" if y == 0 else f"y{y}"
        register(f"bs_total_assets_{label}", "BS", f"{col}5", "output", f"bs_total_assets_{label}")
        register(f"bs_total_lae_{label}", "BS", f"{col}8", "output", f"bs_total_lae_{label}")
        register(f"bs_cash_{label}", "BS", f"{col}2", "output", f"bs_cash_{label}")

    # ---- CF (Cash Flow) -----------------------------------------------------
    ws_cf = wb.create_sheet("CF")
    for i, h in enumerate(is_headers, 1):
        c = ws_cf.cell(row=1, column=i, value=h)
        style_header(c)
    ws_cf.column_dimensions["A"].width = 30
    for col in range(2, len(is_headers) + 1):
        ws_cf.column_dimensions[get_column_letter(col)].width = 14

    cf_labels = [
        "CFO",                  # row 2
        "CFI (capex)",          # row 3
        "CFF",                  # row 4
        "Net change in cash",   # row 5
        "Cumulative cash",      # row 6 — fed back into BS row 2
        "Check (Δcash vs BS)",  # row 7 — must be 0
    ]
    for i, lbl in enumerate(cf_labels, start=2):
        c = ws_cf.cell(row=i, column=1, value=lbl)
        style_label(c)

    # FY-1 anchor: CF columns for FY-1 are 0 — the anchor period only provides
    # opening balances; we do not back-fill its activity.
    for r in range(2, 7):
        ws_cf[f"B{r}"] = 0
    ws_cf["B6"] = "=cash_fy_minus_1"
    ws_cf["B7"] = 0

    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)
        prev = get_column_letter(2 + y - 1)
        # CFO = NI + D&A (add back, D&A is negative on IS) - ΔNWC
        ws_cf[f"{col}2"] = f"=IS!{col}11 - IS!{col}6 - (BS!{col}3 - BS!{prev}3)"
        # CFI = -capex
        ws_cf[f"{col}3"] = f"=-capex_pct_y{y}/100*IS!{col}2"
        # CFF = 0 (skeleton: no dividends, no buybacks, debt flat)
        ws_cf[f"{col}4"] = "=0"
        # net change in cash
        ws_cf[f"{col}5"] = f"={col}2+{col}3+{col}4"
        # cumulative cash
        ws_cf[f"{col}6"] = f"={prev}6+{col}5"
        # tie-out: Δcash from CF vs BS
        ws_cf[f"{col}7"] = f"={col}5-(BS!{col}2-BS!{prev}2)"

    # Register tie-out cells
    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(2 + y)
        register(f"cf_check_y{y}", "CF", f"{col}7", "output", f"cf_check_y{y}")

    # ---- DCF (skeleton, formulas TODO) --------------------------------------
    ws_dcf = wb.create_sheet("DCF")
    c = ws_dcf.cell(row=1, column=1, value="DCF — TODO: formulas stubbed")
    style_header(c)
    ws_dcf.column_dimensions["A"].width = 30
    ws_dcf.column_dimensions["B"].width = 16

    ws_dcf["A3"] = "WACC"
    ws_dcf["B3"] = "=wacc_pct/100"
    style_todo(ws_dcf["B3"])
    register("dcf_wacc", "DCF", "B3", "output", "dcf_wacc")

    ws_dcf["A4"] = "Terminal growth"
    ws_dcf["B4"] = "=terminal_growth_pct/100"
    style_todo(ws_dcf["B4"])
    register("dcf_terminal_growth", "DCF", "B4", "output", "dcf_terminal_growth")

    # Stubbed UFCF row referencing IS so the model is at least non-empty;
    # discount + terminal value formulas are flagged TODO inline.
    ws_dcf["A6"] = "UFCF Y1..Y5 (stub: EBIT*(1-t) + DA - capex - ΔNWC)"
    for y in range(1, FORECAST_YEARS + 1):
        col = get_column_letter(1 + y)  # B..F
        ws_dcf[f"{col}6"] = (
            f"=IS!{get_column_letter(2+y)}7 * (1 - tax_rate_y{y}_pct/100) "
            f"- IS!{get_column_letter(2+y)}6 "
            f"- capex_pct_y{y}/100 * IS!{get_column_letter(2+y)}2 "
            f"- (BS!{get_column_letter(2+y)}3 - BS!{get_column_letter(2+y-1)}3)"
        )
        style_todo(ws_dcf[f"{col}6"])
        register(f"dcf_ufcf_y{y}", "DCF", f"{col}6", "output", f"dcf_ufcf_y{y}")

    ws_dcf["A8"] = "Enterprise value (TODO: discount + TV)"
    ws_dcf["B8"] = '="TODO"'
    style_todo(ws_dcf["B8"])
    register("dcf_enterprise_value", "DCF", "B8", "output", "dcf_enterprise_value")

    # ---- Sensitivity (skeleton, formulas TODO) -------------------------------
    ws_s = wb.create_sheet("Sensitivity")
    c = ws_s.cell(row=1, column=1, value="Sensitivity — WACC × Terminal growth — TODO")
    style_header(c)
    ws_s.column_dimensions["A"].width = 26
    for col in range(2, 8):
        ws_s.column_dimensions[get_column_letter(col)].width = 14

    ws_s["A3"] = "WACC min / max"
    ws_s["B3"] = "=wacc_min_pct/100"
    ws_s["C3"] = "=wacc_max_pct/100"
    style_todo(ws_s["B3"]); style_todo(ws_s["C3"])
    register("sens_wacc_min", "Sensitivity", "B3", "output", "sens_wacc_min")
    register("sens_wacc_max", "Sensitivity", "C3", "output", "sens_wacc_max")

    ws_s["A4"] = "TG min / max"
    ws_s["B4"] = "=tg_min_pct/100"
    ws_s["C4"] = "=tg_max_pct/100"
    style_todo(ws_s["B4"]); style_todo(ws_s["C4"])
    register("sens_tg_min", "Sensitivity", "B4", "output", "sens_tg_min")
    register("sens_tg_max", "Sensitivity", "C4", "output", "sens_tg_max")

    ws_s["A6"] = "Grid (TODO)"
    ws_s["B6"] = '="TODO"'
    style_todo(ws_s["B6"])
    register("sens_grid_anchor", "Sensitivity", "B6", "output", "sens_grid_anchor")

    # ---- Persist -------------------------------------------------------------
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)

    CELL_MAP_PATH.parent.mkdir(parents=True, exist_ok=True)
    CELL_MAP_PATH.write_text(
        json.dumps(cell_map, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    return cell_map


if __name__ == "__main__":
    cm = build()
    print(f"Wrote template:  {TEMPLATE_PATH}")
    print(f"Wrote cell_map:  {CELL_MAP_PATH}")
    inputs = sum(1 for v in cm.values() if v["kind"] == "input")
    outputs = sum(1 for v in cm.values() if v["kind"] == "output")
    print(f"Logical names registered: {len(cm)} ({inputs} inputs, {outputs} outputs)")

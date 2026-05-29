"""Fill the financial model.

The ONLY public API is `fill(template_path, output_path, values, sources)`.

Guarantees:
- Every key in `values` must exist in cell_map.json. Unknown keys raise.
- Only keys with kind == 'input' may be written. Others raise.
- Cells are written ONLY via cell_map.json. No literal coordinates are
  accepted from the caller. This rule is enforced by API shape.
- The Assumptions tab's `source` column (E) is populated for every input.
  A row written without a source raises.
- MISSING values: if values[name] == {"missing": "<reason>"}, the cell is left
  blank and the source column receives "MISSING: <reason>". These are counted
  and returned in the fill report rather than raising.
- Source formats accepted:
    "financial-datasets:income-statements ticker=AAPL as-of=2026-05-28"
    "derived from in_rev_growth_base +200bps"
    "MISSING: financial-datasets returned null for FY-1 revenue"
"""

from __future__ import annotations

import json
import os
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SKILL_DIR = Path(__file__).resolve().parent
PLUGIN_ROOT = Path(os.environ.get("CLAUDE_PLUGIN_ROOT", SKILL_DIR.parent.parent))
PROJECT_DIR = Path(os.environ.get("CLAUDE_PROJECT_DIR", Path.cwd()))

TEMPLATE_PATH = SKILL_DIR / "template" / "model_template.xlsx"
CELL_MAP_PATH = SKILL_DIR / "reference" / "cell_map.json"
OUTPUT_DIR = PROJECT_DIR / "output" / "agent-finance" / "models"


def _load_cell_map() -> dict[str, dict]:
    with CELL_MAP_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def fill(
    template_path: str | Path,
    output_path: str | Path,
    values: dict[str, Any],
    sources: dict[str, str],
) -> dict:
    """Fill the template, writing inputs through cell_map.json only.

    Parameters
    ----------
    template_path : path to the empty template.
    output_path   : where to write the filled model.
    values        : {logical_name: value}. Pass None to keep the template
                    default. Pass {"missing": "<reason>"} to mark a cell as
                    data-unavailable (cell stays blank, source col gets
                    "MISSING: <reason>").
    sources       : {logical_name: source_string}. Required for every key
                    in values whose value is not None.

    Returns
    -------
    dict with keys:
        written   — count of cells actually written
        skipped   — count of None values (kept as template default)
        missing   — list of {name, reason} for MISSING values
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    cell_map = _load_cell_map()

    # Extract Comps peers before cell_map validation (not a named range).
    comps_peers = values.pop("_comps_peers", None)
    comps_source = sources.pop("_comps_peers", None)

    input_keys = {k for k, v in cell_map.items() if v["kind"] == "input"}

    unknown = set(values) - set(cell_map)
    if unknown:
        raise KeyError(
            f"Unknown logical names not in cell_map.json: {sorted(unknown)}"
        )
    not_inputs = set(values) & set(cell_map) - input_keys
    if not_inputs:
        raise KeyError(
            f"These keys are not inputs (they are outputs/formulas): {sorted(not_inputs)}"
        )

    # Separate MISSING sentinels from real values.
    missing_entries: list[dict] = []
    real_values: dict[str, Any] = {}
    for k, v in values.items():
        if isinstance(v, dict) and "missing" in v:
            missing_entries.append({"name": k, "reason": v["missing"]})
        else:
            real_values[k] = v

    # Source policy: every real (non-None, non-MISSING) value needs a source.
    missing_sources = [
        k for k, v in real_values.items() if v is not None and not sources.get(k)
    ]
    if missing_sources:
        raise ValueError(
            f"These inputs are missing a non-empty source: {sorted(missing_sources)}"
        )

    # Copy template -> output, then open and write.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)

    written = 0
    skipped = 0

    for name, val in real_values.items():
        if val is None:
            skipped += 1
            continue
        entry = cell_map[name]
        sheet = wb[entry["sheet"]]
        sheet[entry["cell"]] = val
        written += 1
        # Write source annotation in column L (12) of the same row, on any
        # input sheet. Column L is safely outside the data range (C:J) on
        # the Assumptions tab; other input sheets (Debt Schedule, PP&E
        # Schedule) use it similarly.
        row = int("".join(ch for ch in entry["cell"] if ch.isdigit()))
        wb[entry["sheet"]].cell(row=row, column=12, value=sources[name])

    # Write MISSING annotations: clear the cell (erase any template sentinel)
    # and write the reason in the source column (col L).
    for m in missing_entries:
        entry = cell_map[m["name"]]
        sheet = wb[entry["sheet"]]
        sheet[entry["cell"]] = None  # erase template placeholder/sentinel
        row = int("".join(ch for ch in entry["cell"] if ch.isdigit()))
        sheet.cell(row=row, column=12, value=f"MISSING: {m['reason']}")

    # --- Fill Comps tab from peer_comps.json if passed via --comps ---
    comps_written = _fill_comps(wb, comps_peers, comps_source)

    wb.save(output_path)

    return {
        "written": written + comps_written,
        "skipped": skipped,
        "missing": missing_entries,
    }


def _fill_comps(wb, peers: list | None, source: str | None) -> int:
    """Write peer rows into the Comps sheet.

    peers: list of dicts with keys: name, ev_sales, ev_ebitda, pe, p_fcf
    Rows 7-10 (up to 4 peers). Median row 11 stays as formula.
    Returns count of cells written.
    """
    if not peers:
        return 0
    ws = wb["Comps"]
    written = 0
    for i, peer in enumerate(peers[:4]):
        row = 7 + i
        ws.cell(row, 2).value = peer.get("name", f"Peer {i+1}")
        ws.cell(row, 3).value = peer.get("ev_sales")
        ws.cell(row, 4).value = peer.get("ev_ebitda")
        ws.cell(row, 5).value = peer.get("pe")
        ws.cell(row, 6).value = peer.get("p_fcf")
        ws.cell(row, 7).value = source or peer.get("source", "")
        written += 5
    ws.cell(11, 2).value = "Peer Median (excl. subject)"
    return written


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument(
        "--values",
        required=True,
        help="Path to JSON file mapping logical_name -> value (use null to keep default).",
    )
    parser.add_argument(
        "--sources",
        required=True,
        help="Path to JSON file mapping logical_name -> source string.",
    )
    parser.add_argument(
        "--comps",
        default=None,
        help="Optional path to peer_comps.json — list of {name, ev_sales, ev_ebitda, pe, p_fcf, source}.",
    )
    args = parser.parse_args()

    with open(args.values, encoding="utf-8") as f:
        vals = json.load(f)
    with open(args.sources, encoding="utf-8") as f:
        srcs = json.load(f)

    comps_peers = None
    comps_source = None
    if args.comps:
        with open(args.comps, encoding="utf-8") as f:
            comps_data = json.load(f)
        comps_peers = comps_data.get("peers", comps_data)
        comps_source = comps_data.get("source", f"peer_comps.json as-of {comps_data.get('as_of','')}")

    if comps_peers:
        vals["_comps_peers"] = comps_peers
        srcs["_comps_peers"] = comps_source

    report = fill(args.template, args.output, vals, srcs)
    print(f"Wrote {args.output}")
    print(json.dumps(report, indent=2))

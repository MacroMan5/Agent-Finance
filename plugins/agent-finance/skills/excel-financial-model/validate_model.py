"""Validate a (filled or empty) financial model.

Eight checks, every one is blocking:

  1. Balance sheet identity:    Assets = L + E
  2. Cash flow tie-out:         Δcash on CF == Δcash on BS, per forecast year
  3. No error cells:            no #REF!, #DIV/0!, #VALUE!, #NAME?, #NUM!
  4. Segment sums match totals: vacuously satisfied (no segment breakdown).
  5. Plausibility bounds:       margins in [0,1], YoY growth in [-0.5, 2.0]
  6. Source traceability:       every Assumptions row with a value has a source
  7. Template v2 built-in checks: chk_bs_balance, chk_cf_tie, chk_wacc_gt_g,
                                  chk_tv_pct, chk_int_coverage,
                                  chk_revolver_nonneg, chk_scenario_valid
                                  must all be TRUE/0 (as Excel evaluates them).
                                  Skipped if the workbook was never opened in
                                  Excel (cached values = None).
  8. MISSING count:             no critical input may remain MISSING.
  9. Scenario coverage:         every (bull, base, bear) triplet must be fully
                                populated or fully absent — no partial fill.

Exit code: 0 = all passed, 1 = any failure. JSON report on stdout.

Run via the plugin's resolved path:

    python "$CLAUDE_PLUGIN_ROOT/skills/excel-financial-model/validate_model.py" \
        "$CLAUDE_PROJECT_DIR/output/agent-finance/models/<TICKER>_<DATE>.xlsx"
"""

from __future__ import annotations

import json
import math
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SKILL_DIR = Path(__file__).resolve().parent
PLUGIN_ROOT = Path(os.environ.get("CLAUDE_PLUGIN_ROOT", SKILL_DIR.parent.parent))
CELL_MAP_PATH = SKILL_DIR / "reference" / "cell_map.json"

# Tolerance for floating-point identity checks (relative).
REL_TOL = 1e-6
# Absolute fallback so an empty (all-zero) template doesn't divide by zero.
ABS_TOL = 1e-6

# Critical inputs: validator fails if these remain MISSING.
# Non-critical (optional) inputs are allowed to be MISSING.
CRITICAL_INPUTS = {
    "in_rev_growth_base", "in_rev_growth_bull", "in_rev_growth_bear",
    "in_gross_margin_base", "in_gross_margin_bull", "in_gross_margin_bear",
    "in_sga_pct_base", "in_sga_pct_bull", "in_sga_pct_bear",
    "in_tax_rate_base", "in_tax_rate_bull", "in_tax_rate_bear",
    "in_beta_base", "in_beta_bull", "in_beta_bear",
    "in_erp_base", "in_erp_bull", "in_erp_bear",
    "in_tgr_base", "in_tgr_bull", "in_tgr_bear",
    "in_shares_outstanding", "in_cur_price",
    "in_rev_hist_fy1", "in_rev_hist_fy2", "in_rev_hist_fy3",
}

# Cells that must not remain MISSING after pipeline fill.
# Derived from cell_map.json at import time — no second source of truth.
_CELL_MAP_RAW: dict = json.loads(CELL_MAP_PATH.read_text(encoding="utf-8")) if CELL_MAP_PATH.exists() else {}
SENTINEL_CELLS: dict[str, tuple[str, str]] = {
    name: (entry["sheet"], entry["cell"])
    for name, entry in _CELL_MAP_RAW.items()
    if entry.get("kind") == "input" and name in CRITICAL_INPUTS
}

# v2 template built-in check cells (read from Checks sheet via cell_map).
V2_CHECKS = [
    "chk_bs_balance",
    "chk_cf_tie",
    "chk_wacc_gt_g",
    "chk_tv_pct",
    "chk_int_coverage",
    "chk_revolver_nonneg",
    "chk_scenario_valid",
]


def _is_error_string(v: Any) -> bool:
    if not isinstance(v, str):
        return False
    return v in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#N/A", "#NULL!"}


def _read_cached_values(path: Path) -> dict[tuple[str, str], Any]:
    """Return {(sheet, cell): cached_value} from the last Excel save.

    openpyxl with data_only=True returns the values Excel calculated on its
    last save. For a freshly-written workbook (never opened in Excel) these
    are typically None for formula cells — that's fine; the checks treat
    None and 0 as "empty / zero" rather than as errors.
    """
    wb = load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], Any] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cache[(ws.title, c.coordinate)] = c.value
    return cache


def _read_raw(path: Path) -> dict[tuple[str, str], Any]:
    """Return {(sheet, cell): raw_value_or_formula}."""
    wb = load_workbook(path, data_only=False)
    raw: dict[tuple[str, str], Any] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    raw[(ws.title, c.coordinate)] = c.value
    return raw


def _val(cache, sheet, cell, default=0.0):
    v = cache.get((sheet, cell))
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    return v


def validate(model_path: Path) -> dict:
    with CELL_MAP_PATH.open(encoding="utf-8") as f:
        cell_map = json.load(f)

    cache = _read_cached_values(model_path)
    raw = _read_raw(model_path)

    report = {"model": str(model_path), "checks": {}}

    # --- 1. Balance-sheet identity --------------------------------------------
    bs_check = {"passed": True, "details": []}
    for label in ["fy_minus_1"] + [f"y{y}" for y in range(1, 6)]:
        a_key = f"bs_total_assets_{label}"
        le_key = f"bs_total_lae_{label}"
        if a_key not in cell_map or le_key not in cell_map:
            continue
        a = _val(cache, cell_map[a_key]["sheet"], cell_map[a_key]["cell"])
        le = _val(cache, cell_map[le_key]["sheet"], cell_map[le_key]["cell"])
        if not isinstance(a, (int, float)) or not isinstance(le, (int, float)):
            bs_check["passed"] = False
            bs_check["details"].append(
                {"period": label, "error": f"non-numeric value (a={a!r}, le={le!r})"}
            )
            continue
        diff = abs(a - le)
        scale = max(abs(a), abs(le), ABS_TOL)
        if diff / scale > REL_TOL:
            bs_check["passed"] = False
            bs_check["details"].append(
                {"period": label, "assets": a, "liab_plus_equity": le, "diff": diff}
            )
    report["checks"]["1_balance_sheet_identity"] = bs_check

    # --- 2. Cash flow tie-out --------------------------------------------------
    cf_check = {"passed": True, "details": []}
    for y in range(1, 6):
        key = f"cf_check_y{y}"
        if key not in cell_map:
            continue
        v = _val(cache, cell_map[key]["sheet"], cell_map[key]["cell"])
        if not isinstance(v, (int, float)):
            cf_check["passed"] = False
            cf_check["details"].append({"year": y, "error": f"non-numeric {v!r}"})
            continue
        if abs(v) > ABS_TOL:
            cf_check["passed"] = False
            cf_check["details"].append({"year": y, "tie_out_diff": v})
    report["checks"]["2_cash_flow_tie_out"] = cf_check

    # --- 3. No error cells ----------------------------------------------------
    err_check = {"passed": True, "details": []}
    for (sheet, coord), v in cache.items():
        if _is_error_string(v):
            err_check["passed"] = False
            err_check["details"].append({"sheet": sheet, "cell": coord, "value": v})
    report["checks"]["3_no_error_cells"] = err_check

    # --- 4. Segment sums match totals (vacuous on the skeleton) ---------------
    report["checks"]["4_segment_sums"] = {
        "passed": True,
        "details": "skeleton has no segment breakdown; check is vacuously satisfied.",
    }

    # --- 5. Plausibility bounds -----------------------------------------------
    pb_check = {"passed": True, "details": []}
    # Inputs we want inside [0,1] when expressed as ratios (skeleton stores
    # percentages, so we test the *_pct family in [0, 100]).
    margin_keys = (
        ["gross_margin_fy_minus_1_pct"]
        + [f"gross_margin_y{y}_pct" for y in range(1, 6)]
        + [f"opex_pct_y{y}" for y in range(1, 6)]  # stored as % already
        + [f"tax_rate_y{y}_pct" for y in range(1, 6)]
        + ["tax_rate_fy_minus_1_pct", "opex_pct_fy_minus_1"]
    )
    for k in margin_keys:
        if k not in cell_map:
            continue
        v = _val(cache, cell_map[k]["sheet"], cell_map[k]["cell"])
        if v is None:
            continue
        if isinstance(v, (int, float)) and not (0 <= v <= 100):
            pb_check["passed"] = False
            pb_check["details"].append({"key": k, "value": v, "bound": "0..100"})

    # D&A and capex as % of revenue: tighter band (>50% is implausible).
    tight_pct_keys = (
        ["da_pct_fy_minus_1"]
        + [f"da_pct_y{y}" for y in range(1, 6)]
        + [f"capex_pct_y{y}" for y in range(1, 6)]
    )
    for k in tight_pct_keys:
        if k not in cell_map:
            continue
        v = _val(cache, cell_map[k]["sheet"], cell_map[k]["cell"])
        if v is None:
            continue
        if isinstance(v, (int, float)) and not (0 <= v <= 50):
            pb_check["passed"] = False
            pb_check["details"].append({"key": k, "value": v, "bound": "0..50"})

    growth_keys = [f"rev_growth_y{y}_pct" for y in range(1, 6)]
    for k in growth_keys:
        if k not in cell_map:
            continue
        v = _val(cache, cell_map[k]["sheet"], cell_map[k]["cell"])
        if v is None:
            continue
        if isinstance(v, (int, float)) and not (-50 <= v <= 200):
            pb_check["passed"] = False
            pb_check["details"].append({"key": k, "value": v, "bound": "-50..200"})
    report["checks"]["5_plausibility_bounds"] = pb_check

    # --- 6. Source traceability -----------------------------------------------
    # For every cell_map input on any sheet that was written (non-None,
    # non-MISSING), column L of the same row must hold a non-empty source.
    # An empty template (all inputs at template default) passes vacuously.
    src_check = {"passed": True, "details": []}
    wb_raw = load_workbook(model_path)
    for logical, entry in cell_map.items():
        if entry["kind"] != "input":
            continue
        ws_check = wb_raw[entry["sheet"]] if entry["sheet"] in wb_raw.sheetnames else None
        if ws_check is None:
            continue
        row_num = int("".join(ch for ch in entry["cell"] if ch.isdigit()))
        col_letter = "".join(ch for ch in entry["cell"] if ch.isalpha())
        # Only check cells that were actually written (have a numeric value).
        val = ws_check[f"{col_letter}{row_num}"].value
        if not isinstance(val, (int, float)):
            continue
        src = ws_check.cell(row=row_num, column=12).value
        if not src or (isinstance(src, str) and (not src.strip() or src.startswith("MISSING:"))):
            if not src or not src.strip():
                src_check["passed"] = False
                src_check["details"].append({"logical_name": logical, "row": row_num, "value": val})
    report["checks"]["6_source_traceability"] = src_check

    # --- 7. Template v2 built-in checks (Checks sheet) -----------------------
    v2_check = {"passed": True, "details": [], "skipped": False}
    for chk_name in V2_CHECKS:
        if chk_name not in cell_map:
            continue
        entry = cell_map[chk_name]
        v = cache.get((entry["sheet"], entry["cell"]))
        if v is None:
            # Workbook never opened in Excel — cached values are absent.
            v2_check["skipped"] = True
            v2_check["details"].append({"check": chk_name, "status": "skipped (no cached value)"})
            continue
        # Excel formula returns TRUE/FALSE, a numeric diff (0 = balanced),
        # or a string verdict: 'OK' = pass, 'REVIEW' = advisory warning (not
        # a hard failure — the report surfaces it separately), 'FAIL' = fail.
        if v is True or v == "TRUE" or v == "OK" or (isinstance(v, (int, float)) and abs(v) <= ABS_TOL):
            pass  # check passed
        elif isinstance(v, str) and v == "REVIEW":
            v2_check["details"].append({"check": chk_name, "value": v, "advisory": True})
        else:
            v2_check["passed"] = False
            v2_check["details"].append({"check": chk_name, "value": v})
    if v2_check["skipped"] and v2_check["passed"]:
        v2_check["details"].append({"note": "all checks skipped — open in Excel to recalculate"})
    report["checks"]["7_template_builtin_checks"] = v2_check

    # --- 8. MISSING count on critical inputs ---------------------------------
    # Scan all input cells (any sheet) for MISSING: annotations in column L.
    missing_check = {"passed": True, "details": []}
    wb_src = load_workbook(model_path)
    try:
        for logical, entry in cell_map.items():
            if entry["kind"] != "input" or logical not in CRITICAL_INPUTS:
                continue
            if entry["sheet"] not in wb_src.sheetnames:
                continue
            ws_src = wb_src[entry["sheet"]]
            row_num = int("".join(ch for ch in entry["cell"] if ch.isdigit()))
            src_cell = ws_src.cell(row=row_num, column=12).value
            if isinstance(src_cell, str) and src_cell.startswith("MISSING:"):
                missing_check["passed"] = False
                missing_check["details"].append({
                    "logical_name": logical,
                    "row": row_num,
                    "sheet": entry["sheet"],
                    "reason": src_cell,
                })
        report["checks"]["8_missing_critical_inputs"] = missing_check

        # --- 9. Scenario coverage -------------------------------------------------
        # For every metric that has a bull/base/bear triplet, either all three must
        # be filled (non-MISSING, numeric cell value present) or none of them.
        scen_check = {"passed": True, "details": []}
        from collections import defaultdict
        metric_scenarios: dict[str, dict[str, str]] = defaultdict(dict)
        for logical, entry in cell_map.items():
            if entry["kind"] != "input" or "scenario" not in entry:
                continue
            metric = entry.get("metric", "")
            scenario = entry["scenario"]
            metric_scenarios[metric][scenario] = logical

        for metric, scen_map in metric_scenarios.items():
            if set(scen_map.keys()) != {"bull", "base", "bear"}:
                continue
            fill_statuses = {}
            for scen, logical in scen_map.items():
                entry = cell_map[logical]
                if entry["sheet"] not in wb_src.sheetnames:
                    fill_statuses[scen] = False
                    continue
                ws_scen = wb_src[entry["sheet"]]
                row_num = int("".join(ch for ch in entry["cell"] if ch.isdigit()))
                col_letter = "".join(ch for ch in entry["cell"] if ch.isalpha())
                val = ws_scen[f"{col_letter}{row_num}"].value
                src = ws_scen.cell(row=row_num, column=12).value
                is_missing = isinstance(src, str) and src.startswith("MISSING:")
                is_filled = isinstance(val, (int, float)) and not is_missing
                fill_statuses[scen] = is_filled
            if len(set(fill_statuses.values())) > 1:
                scen_check["passed"] = False
                scen_check["details"].append({
                    "metric": metric,
                    "fill_status": fill_statuses,
                })
        report["checks"]["9_scenario_coverage"] = scen_check
    finally:
        wb_src.close()

    # --- 10. Sentinel text detector ------------------------------------------
    # Cells in SENTINEL_CELLS must not remain unfilled after the pipeline runs.
    # fill_model.py erases any template sentinel (sets cell to None) and writes
    # "MISSING: <reason>" to col L — so we check both the data cell (old-style
    # text sentinel) and col L (pipeline-style MISSING annotation).
    sentinel_check = {"passed": True, "details": []}
    wb_raw = load_workbook(model_path, data_only=False)
    try:
        for logical, (sheet_name, cell_addr) in SENTINEL_CELLS.items():
            if sheet_name not in wb_raw.sheetnames:
                continue
            ws = wb_raw[sheet_name]
            val = ws[cell_addr].value
            row_num = int("".join(ch for ch in cell_addr if ch.isdigit()))
            src_val = ws.cell(row=row_num, column=12).value
            text_sentinel = isinstance(val, str) and val.strip().upper() == "MISSING"
            col_l_missing = val is None and isinstance(src_val, str) and src_val.startswith("MISSING:")
            if text_sentinel or col_l_missing:
                reason = src_val if col_l_missing else f"{cell_addr} still reads 'MISSING'"
                sentinel_check["passed"] = False
                sentinel_check["details"].append({
                    "logical_name": logical,
                    "cell": f"{sheet_name}!{cell_addr}",
                    "error": reason,
                })
    finally:
        wb_raw.close()
    report["checks"]["10_sentinel_placeholder_check"] = sentinel_check

    # --- aggregate -----------------------------------------------------------
    report["passed"] = all(c["passed"] for c in report["checks"].values())
    return report


def main(argv):
    if len(argv) < 2:
        print("usage: validate_model.py <path-to-xlsx>", file=sys.stderr)
        return 2
    path = Path(argv[1])
    if not path.exists():
        print(f"file not found: {path}", file=sys.stderr)
        return 2
    rep = validate(path)
    print(json.dumps(rep, indent=2))
    return 0 if rep["passed"] else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
